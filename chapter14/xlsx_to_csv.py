#! /usr/bin/python3

# xlsx_to_csv.py - Program to convert files in current directory from .xlsx to .csv

# Usage: ./xlsx_to_csv.py 

# Import modules
import os, sys, openpyxl, csv
from openpyxl.utils import get_column_letter as g_c_l

# Create csv subfolder in which new .csv files will be created
dest = 'csv'
os.makedirs(dest, exist_ok = True)

# Loop through files in current directory
for filename in os.listdir():

	# Check file is type .xls or .xlsx
	if not filename.lower().endswith('.xlsx'):
		continue

	# Open spreadsheet file
	f = open(filename, 'rb')
	wb = openpyxl.load_workbook(f)

	# Loop through sheets
	for sheetname in wb.get_sheet_names():
		sheet = wb.get_sheet_by_name(sheetname)

		# Get max dimensions of sheet contents
		rows = sheet.max_row
		cols = sheet.max_column

		# Open new .csv file and create csv.writer object
		(basename, extension) = os.path.splitext(filename)
		new_filename = basename + '_' + sheetname + '.csv'
		path = os.path.join(dest, new_filename)
		f = open(path, 'w', newline = '')
		csv_file = csv.writer(f)

		# Loop through rows
		for row in range(1, rows + 1):

			# Create new line
			line = [] 

			# Loop through columns
			for col in range(1, cols + 1):

				cell = g_c_l(col) + str(row) # Generate cell reference string
				line.append(sheet[cell].value) # Append cell values to the line

			# Write list object to .csv file
			csv_file.writerow(line)

		# Close the newly written .csv file
		f.close()
