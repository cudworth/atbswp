#! /usr/bin/python3
# spreadsheet_cell_inverter.py
# Program to transpose a spreadsheet
# Usage: ./spreadsheet_cell_inverter.py <filename>

# Import modules
import sys, os, openpyxl
from openpyxl.utils import get_column_letter as g_c_l

# Import filename from sys.argv
try:
	filename = sys.argv[1]
except:
	print('Improper usage')
	exit()

assert os.path.exists(filename), 'File does not exist'

# Open filename
w1 = openpyxl.load_workbook(filename)
s1 = w1.active

# Open new spreadsheet
w2 = openpyxl.Workbook()
s2 = w2.active

# Size of contents in original file
rows = s1.max_row
cols = s1.max_column

# Read in contents of spreadsheet
for row in range(1, rows + 1):

	for col in range(1, cols + 1):

		orig = g_c_l(col) + str(row)
		dest = g_c_l(row) + str(col)
		s2[dest] = s1[orig].value

# Write new spreadsheet object
w2.save('_' + filename)

# Close spreadsheet objects
w1.close()
w2.close()
