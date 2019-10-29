#! /usr/bin/python3

# Program to insert blank rows into an Excel file

# Usage: blank_row_inserter.py <N> <M> <filename.xlsx>
# Inserts M blank rows at row N in the existing file provided

# Import modules
import sys, os, openpyxl
from openpyxl.utils import get_column_letter as g_c_l

# Import parameters from sys.argv
try:
	M = int(sys.argv[1])
	N = int(sys.argv[2])
	filename = sys.argv[3]

except:
	print('Improper usage')
	exit()

assert os.path.exists(filename),'File does not exist'

# Open workbook for editing
w1 = openpyxl.load_workbook(filename)
s1 = w1.active

# Create new workbook
w2 = openpyxl.Workbook()
s2 = w2.active

# Dimensions of original workbook
rows = s1.max_row
cols = s1.max_column 
# Create list of column letters
cl = []
for col in range(1, cols + 1):
	cl.append(g_c_l(col))

for row in range(1, M):

	for col in range(cols):
		orig = dest = cl[col] + str(row)

		s2[dest] = s1[orig].value

for row in range(M, rows + 1):

	for col in range(cols):
		orig = cl[col] + str(row) # Origin cell
		dest = cl[col] + str(row + N) # Destination cell

		s2[dest] = s1[orig].value
		
w1.close()
w2.save('_' + filename)
