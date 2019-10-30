#! /usr/bin/python3

# spreadsheet_to_text_files.py

# A program to read user specified spreadsheet and write individual text files containing the contents of sheet columns.

# Usage: spreadsheet_to_text_files.py <filename> 

# Import modules
import os, sys, openpyxl
from openpyxl.utils import get_column_letter as g_c_l

# Parse inputs
try:
	filename = sys.argv[1]
except:
	print('Improper usage')
	exit()

# Verify filename provided
assert os.path.exists(filename), 'File %s could not be found' % filename

# Open a spreadsheet
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

# Find dims of spreadsheet content
rows = sheet.max_row
cols = sheet.max_column

# Loop through columns
for col in range(1, cols + 1):

	cl = g_c_l(col) # Find column letter
	newfile = 'column_' + cl + '.txt' # Create new filename
	f = open(newfile,'w') # Create new file w/ filename	

	# Loop through rows
	for row in range(1, rows + 1):
		cell = cl + str(row)
		line = sheet[cell].value
		if line == None:
			line = ''
		f.write(line)

	# Close file
	f.close()

# Close the workbook
workbook.close()




	
